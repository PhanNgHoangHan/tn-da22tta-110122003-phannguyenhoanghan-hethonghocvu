import csv
import io
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponse
import openpyxl
from .models import SinhVien, MonHoc, HocKy, Nganh
from .forms import SinhVienForm, MonHocForm, HocKyForm, NganhForm, ImportCSVForm


def role_required(*roles):
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('accounts:login')
            if request.user.role not in roles and not request.user.is_admin:
                messages.error(request, 'Bạn không có quyền truy cập.')
                return redirect('dashboard:index')
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


# ---- Sinh viên ----
@login_required
def sinhvien_list(request):
    qs = SinhVien.objects.select_related('nganh', 'covan', 'lop').order_by('khoa', 'nganh__khoa', 'lop__ten_lop', 'mssv')
    q = request.GET.get('q', '')
    trang_thai = request.GET.get('trang_thai', '')
    
    # Lấy các bộ lọc phân cấp
    khoa = request.GET.get('khoa', '')
    nganh_id = request.GET.get('nganh', '')
    lop_id = request.GET.get('lop', '')
    khoa_hoc = request.GET.get('khoa_hoc', '')

    if request.user.is_sinhvien:
        try:
            qs = qs.filter(user=request.user)
        except Exception:
            qs = qs.none()
    else:
        if request.user.is_covan:
            qs = qs.filter(lop__covan=request.user)
        
        # Áp dụng bộ lọc phân cấp cho cả Giáo vụ và Cố vấn
        if khoa_hoc:
            cohort_suffix = khoa_hoc[-2:]
            qs = qs.filter(lop__ten_lop__contains=cohort_suffix)
        if khoa:
            qs = qs.filter(nganh__khoa=khoa)
        if nganh_id:
            qs = qs.filter(nganh_id=nganh_id)
        if lop_id:
            qs = qs.filter(lop_id=lop_id)

    if trang_thai:
        qs = qs.filter(trang_thai=trang_thai)

    if q:
        from results.utils import remove_accents
        q_clean = remove_accents(q)
        qs = [
            sv for sv in qs
            if q_clean in remove_accents(sv.mssv)
            or q_clean in remove_accents(sv.ho_ten)
            or (sv.lop and q_clean in remove_accents(sv.lop.ten_lop))
        ]

    # Dữ liệu cho bộ lọc
    from .models import Lop
    import re
    
    nganhs = Nganh.objects.exclude(ma_nganh__iexact='None').exclude(ten_nganh__iexact='None')
    khoas = [k for k in Nganh.objects.exclude(khoa__isnull=True).exclude(khoa='').values_list('khoa', flat=True).distinct() if k.lower() != 'none']
    
    # Lấy danh sách khóa học duy nhất từ các lớp
    unique_years = set()
    for name in Lop.objects.values_list('ten_lop', flat=True):
        if name:
            m = re.search(r'\d{2}', name)
            if m:
                unique_years.add("20" + m.group())
    khoa_hocs = sorted(list(unique_years), reverse=True)

    lops = Lop.objects.exclude(ten_lop__iexact='None')
    if request.user.is_covan:
        lops = lops.filter(covan=request.user)
    if nganh_id:
        lops = lops.filter(nganh_id=nganh_id)
    if khoa_hoc:
        cohort_suffix = khoa_hoc[-2:]
        lops = lops.filter(ten_lop__contains=cohort_suffix)
    
    if khoa:
        nganhs = nganhs.filter(khoa=khoa)

    return render(request, 'students/sinhvien_list.html', {
        'sinhviens': qs, 
        'nganhs': nganhs, 
        'khoas': khoas,
        'lops': lops,
        'khoa_hocs': khoa_hocs,
        'q': q,
        'trang_thai': trang_thai, 
        'selected_khoa': khoa,
        'nganh_id': nganh_id,
        'selected_lop': lop_id,
        'selected_khoa_hoc': khoa_hoc,
    })


@login_required
def sinhvien_detail(request, pk):
    sv = get_object_or_404(SinhVien, pk=pk)
    if request.user.is_sinhvien:
        linked_sv = SinhVien.objects.filter(user=request.user).first()
        if linked_sv is None or linked_sv.pk != sv.pk:
            messages.error(request, 'Bạn không có quyền xem thông tin này.')
            return redirect('dashboard:index')

    from results.models import KetQuaHocTap
    from academic_warnings.models import CanhBaoHocVu
    from results.utils import tinh_dtbctl, la_dat, diem_he4, la_gdtc
    from collections import OrderedDict

    canh_baos = CanhBaoHocVu.objects.filter(sinh_vien=sv).order_by('-ngay_tao')

    # Chỉ lấy lần học 1 để hiển thị (lần 2 chỉ dùng cho tích lũy)
    # Sắp xếp: năm giảm dần, HK giảm dần, môn tăng dần
    ket_qua_all = KetQuaHocTap.objects.filter(
        sinh_vien=sv, lan_hoc=1
    ).select_related('mon_hoc', 'hoc_ky').order_by(
        '-hoc_ky__nam_hoc', '-hoc_ky__ky', 'mon_hoc__ten_mh'
    )

    ket_qua_theo_nam = OrderedDict()
    for kq in ket_qua_all:
        nam = kq.hoc_ky.nam_hoc
        hk_ten = str(kq.hoc_ky)
        if nam not in ket_qua_theo_nam:
            ket_qua_theo_nam[nam] = OrderedDict()
        if hk_ten not in ket_qua_theo_nam[nam]:
            ket_qua_theo_nam[nam][hk_ten] = {
                'ket_qua': [], 'tong_mon': 0,
                'tc_dang_ky': 0, 'tc_dat': 0,
                'dtbchk': 0.0, 'dtbchk_4': 0.0
            }
        ket_qua_theo_nam[nam][hk_ten]['ket_qua'].append(kq)

    # Tính thống kê từng HK (chỉ dựa trên lan_hoc=1)
    for nam, hk_dict in ket_qua_theo_nam.items():
        for hk_ten, data in hk_dict.items():
            kqs = data['ket_qua']
            data['tong_mon'] = len(kqs)
            data['tc_dang_ky'] = sum(k.mon_hoc.so_tc for k in kqs)
            data['tc_dat'] = sum(k.mon_hoc.so_tc for k in kqs if la_dat(k.diem_tk))
            
            # Loại bỏ môn học Giáo dục thể chất khi tính GPA học kỳ
            tong_tc = sum(k.mon_hoc.so_tc for k in kqs if k.diem_tk is not None and not la_gdtc(k.mon_hoc.ma_mh))
            tong_d10 = sum(k.diem_tk * k.mon_hoc.so_tc for k in kqs if k.diem_tk is not None and not la_gdtc(k.mon_hoc.ma_mh))
            tong_d4  = sum((diem_he4(k.diem_tk) or 0) * k.mon_hoc.so_tc for k in kqs if k.diem_tk is not None and not la_gdtc(k.mon_hoc.ma_mh))
            
            data['dtbchk']   = round(tong_d10 / tong_tc, 2) if tong_tc > 0 else 0.0
            data['dtbchk_4'] = round(tong_d4  / tong_tc, 2) if tong_tc > 0 else 0.0
            
            # Tính điểm tích lũy tính đến học kỳ này
            hk_obj = kqs[0].hoc_ky if kqs else None
            if hk_obj:
                ctl_10, ctl_4, tc_tl_val, _ = tinh_dtbctl(sv, hk_obj)
            else:
                ctl_10 = ctl_4 = 0.0
                tc_tl_val = 0
            
            data['dtbctl_10'] = ctl_10
            data['dtbctl_4']  = ctl_4
            data['tc_tl']     = tc_tl_val

    # TC tích lũy (tính từ điểm tốt nhất mỗi môn, kể cả lần 2)
    dtbctl_10, dtbctl_4, tc_tl, tc_da_hoc = tinh_dtbctl(sv)

    # Tổng kết toàn khóa (sử dụng tinh_dtbctl để lấy điểm tích lũy chính xác)
    toan_khoa_dtb_10 = dtbctl_10
    toan_khoa_dtb_4  = dtbctl_4

    return render(request, 'students/sinhvien_detail.html', {
        'sv': sv,
        'canh_baos': canh_baos,
        'ket_qua_theo_nam': ket_qua_theo_nam,
        'toan_khoa_dtb_10': toan_khoa_dtb_10,
        'toan_khoa_dtb_4': toan_khoa_dtb_4,
        'tc_tl': tc_tl,
        'tc_da_hoc': tc_da_hoc,
    })


@login_required
@role_required('giaovu', 'admin')
def sinhvien_create(request):
    form = SinhVienForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Thêm sinh viên thành công.')
        return redirect('students:sinhvien_list')
    return render(request, 'students/sinhvien_form.html', {'form': form, 'title': 'Thêm sinh viên'})


@login_required
@role_required('giaovu', 'admin')
def sinhvien_edit(request, pk):
    messages.error(request, 'Chức năng sửa thông tin sinh viên đã bị vô hiệu hóa.')
    return redirect('students:sinhvien_list')


@login_required
@role_required('giaovu', 'admin')
def sinhvien_delete(request, pk):
    messages.error(request, 'Chức năng xóa sinh viên đã bị vô hiệu hóa.')
    return redirect('students:sinhvien_list')


# ---- Môn học ----
@login_required
def monhoc_list(request):
    from collections import OrderedDict
    nganhs = Nganh.objects.all()
    nganh_id = request.GET.get('nganh', '')
    
    if not nganh_id:
        qs = MonHoc.objects.none()
    else:
        qs = MonHoc.objects.select_related('nganh').filter(nganh_id=nganh_id).order_by('-nam_hoc_ctdt', '-hoc_ky_ctdt', 'ma_mh')
        
    q = request.GET.get('q', '')
    if q:
        from results.utils import remove_accents
        q_clean = remove_accents(q)
        qs = [
            mh for mh in qs
            if q_clean in remove_accents(mh.ma_mh)
            or q_clean in remove_accents(mh.ten_mh)
        ]

    # Nhóm môn học theo Học kỳ (8 giảm dần về 1 và None)
    grouped_data = OrderedDict()
    for hk_num in range(8, 0, -1):
        grouped_data[hk_num] = []
    grouped_data[None] = []

    for mh in qs:
        y = mh.nam_hoc_ctdt
        s = mh.hoc_ky_ctdt
        if y and s and 1 <= y <= 4 and 1 <= s <= 2:
            hk_num = (y - 1) * 2 + s
        else:
            hk_num = None
        grouped_data[hk_num].append(mh)

    # Loại bỏ các nhóm rỗng
    final_grouped_data = OrderedDict()
    for hk_num, subjects in grouped_data.items():
        if len(subjects) > 0:
            final_grouped_data[hk_num] = subjects

    return render(request, 'students/monhoc_list.html', {
        'grouped_monhocs': final_grouped_data, 
        'q': q,
        'nganhs': nganhs,
        'selected_nganh': nganh_id
    })


@login_required
@role_required('giaovu', 'admin')
def monhoc_create(request):
    form = MonHocForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Thêm môn học thành công.')
        return redirect('students:monhoc_list')
    return render(request, 'students/monhoc_form.html', {'form': form, 'title': 'Thêm chương trình đào tạo'})


@login_required
@role_required('giaovu', 'admin')
def monhoc_edit(request, pk):
    messages.error(request, 'Chức năng sửa môn học đã bị vô hiệu hóa.')
    return redirect('students:monhoc_list')


@login_required
@role_required('giaovu', 'admin')
def monhoc_delete(request, pk):
    messages.error(request, 'Chức năng xóa môn học đã bị vô hiệu hóa.')
    return redirect('students:monhoc_list')


# ---- Học kỳ ----
@login_required
def hocky_list(request):
    hockys = HocKy.objects.all()
    return render(request, 'students/hocky_list.html', {'hockys': hockys})


@login_required
@role_required('giaovu', 'admin')
def hocky_create(request):
    form = HocKyForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Thêm học kỳ thành công.')
        return redirect('students:hocky_list')
    return render(request, 'students/hocky_form.html', {'form': form, 'title': 'Thêm học kỳ'})


@login_required
@role_required('giaovu', 'admin')
def hocky_edit(request, pk):
    messages.error(request, 'Chức năng sửa học kỳ đã bị vô hiệu hóa.')
    return redirect('students:hocky_list')


# ---- Import CSV ----
def parse_date(date_val):
    from datetime import datetime, date
    if not date_val:
        return None
    if isinstance(date_val, (datetime, date)):
        return date_val
    date_str = str(date_val).strip()
    if date_str.lower() in ('none', ''):
        return None
    
    formats = [
        '%d/%m/%Y',
        '%Y-%m-%d',
        '%d-%m-%Y',
        '%m/%d/%Y',
        '%d/%m/%y',
        '%Y/%m/%d',
    ]
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


@login_required
@role_required('giaovu', 'admin')
def import_sinhvien(request):
    form = ImportCSVForm(request.POST or None, request.FILES or None)
    errors = []
    if request.method == 'POST' and form.is_valid():
        f = request.FILES['file']
        name = f.name.lower()
        count = 0
        try:
            if name.endswith('.csv'):
                decoded = f.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded))
                rows_data = list(reader)
            elif name.endswith(('.xlsx', '.xls')):
                wb = openpyxl.load_workbook(f)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                rows_data = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]
            else:
                messages.error(request, 'Định dạng file không hỗ trợ.')
                return render(request, 'students/import.html', {'form': form, 'title': 'Import sinh viên', 'errors': errors})

            from django.contrib.auth import get_user_model
            User = get_user_model()

            for row in rows_data:
                if not row:
                    continue
                mssv_val = row.get('mssv')
                if mssv_val is None:
                    continue
                mssv = str(mssv_val).strip()
                if not mssv or mssv.lower() == 'none':
                    continue
                try:
                    ma_nganh_val = row.get('ma_nganh')
                    ma_nganh = str(ma_nganh_val).strip() if ma_nganh_val is not None else 'KT'
                    if not ma_nganh or ma_nganh.lower() == 'none':
                        ma_nganh = 'KT'

                    ten_nganh_val = row.get('ten_nganh')
                    ten_nganh = str(ten_nganh_val).strip() if ten_nganh_val is not None else 'Chưa xác định'
                    if not ten_nganh or ten_nganh.lower() == 'none':
                        ten_nganh = 'Chưa xác định'

                    ho_ten_val = row.get('ho_ten')
                    ho_ten = str(ho_ten_val).strip() if ho_ten_val is not None else ''
                    if not ho_ten or ho_ten.lower() == 'none':
                        continue

                    email_val = row.get('email')
                    email = str(email_val).strip() if email_val is not None else ''
                    if email.lower() == 'none':
                        email = ''

                    # gioi_tinh
                    gioi_tinh_val = row.get('gioi_tinh') or row.get('giới tính') or row.get('gioitinh') or row.get('gender')
                    gioi_tinh = ''
                    if gioi_tinh_val:
                        gt_str = str(gioi_tinh_val).strip().lower()
                        if gt_str in ('nam', 'm', 'male'):
                            gioi_tinh = 'Nam'
                        elif gt_str in ('nữ', 'nu', 'female', 'f', 'nữ '):
                            gioi_tinh = 'Nu'

                    # ngay_sinh
                    ngay_sinh_val = row.get('ngay_sinh') or row.get('ngày sinh') or row.get('ngaysinh') or row.get('ngày_sinh') or row.get('dob')
                    ngay_sinh = parse_date(ngay_sinh_val)

                    # khoa ở đây là Khóa học (ví dụ: K2021)
                    khoa_val = row.get('khoa')
                    khoa = str(khoa_val).strip() if khoa_val is not None else ''
                    if khoa.lower() == 'none':
                        khoa = ''
                    if khoa.startswith('K') or khoa.startswith('k'):
                        khoa = khoa[1:]

                    # khoa_vien/ten_khoa ở đây là Khoa chuyên ngành (ví dụ: Khoa Công nghệ thông tin)
                    ten_khoa_val = row.get('ten_khoa') or row.get('khoa_vien') or row.get('khoa_chuyen_nganh') or row.get('khoa_nganh')
                    ten_khoa = str(ten_khoa_val).strip() if ten_khoa_val is not None else 'Khoa Công nghệ thông tin'
                    if not ten_khoa or ten_khoa.lower() == 'none':
                        ten_khoa = 'Khoa Công nghệ thông tin'

                    trang_thai_val = row.get('trang_thai')
                    trang_thai_raw = str(trang_thai_val).strip().lower() if trang_thai_val is not None else ''
                    
                    trang_thai_map = {
                        'dang_hoc': 'dang_hoc',
                        'đang học': 'dang_hoc',
                        'canh_bao': 'canh_bao',
                        'cảnh báo học vụ': 'canh_bao',
                        'cảnh báo': 'canh_bao',
                        'dinh_chi': 'dinh_chi',
                        'đình chỉ': 'dinh_chi',
                        'tot_nghiep': 'tot_nghiep',
                        'tốt nghiệp': 'tot_nghiep',
                        'thoi_hoc': 'thoi_hoc',
                        'thôi học': 'thoi_hoc',
                    }
                    trang_thai = trang_thai_map.get(trang_thai_raw, 'dang_hoc')

                    nganh, created = Nganh.objects.get_or_create(
                        ma_nganh=ma_nganh,
                        defaults={
                            'ten_nganh': ten_nganh,
                            'khoa': ten_khoa
                        }
                    )
                    # Nếu ngành học đã tồn tại nhưng trong file import có tên khoa hoặc tên ngành khác, ta cập nhật lại
                    if not created:
                        updated = False
                        if ten_khoa_val and nganh.khoa != ten_khoa:
                            nganh.khoa = ten_khoa
                            updated = True
                        if ten_nganh_val and ten_nganh != 'Chưa xác định' and nganh.ten_nganh != ten_nganh:
                            nganh.ten_nganh = ten_nganh
                            updated = True
                        if updated:
                            nganh.save()
                    # Tự tạo tài khoản: username = mssv, password = mssv
                    sv_user, user_created = User.objects.get_or_create(
                        username=mssv,
                        defaults={
                            'full_name': ho_ten,
                            'email': email,
                            'role': 'sinhvien',
                        }
                    )
                    if user_created:
                        sv_user.set_password(mssv)
                        sv_user.save()

                    # Tìm hoặc tạo tài khoản cố vấn/giảng viên nếu có
                    ma_gv_val = (
                        row.get('ma_gvcv') or row.get('mã gvcv') or row.get('ma gvcv') or
                        row.get('ma_gv') or row.get('mã gv') or row.get('ma gv') or
                        row.get('ma_covan') or row.get('mã cố vấn') or
                        row.get('mã giảng viên') or row.get('ma_giangvien')
                    )
                    covan = None
                    if ma_gv_val:
                        ma_gv = str(ma_gv_val).strip()
                        if ma_gv and ma_gv.lower() != 'none':
                            ten_gv_val = (
                                row.get('ten_gvcv') or row.get('tên gvcv') or row.get('ten gvcv') or
                                row.get('ten_gv') or row.get('tên gv') or row.get('ten gv') or
                                row.get('ten_covan') or row.get('tên cố vấn') or
                                row.get('tên giảng viên') or row.get('ten_giangvien')
                            )
                            ten_gv = str(ten_gv_val).strip() if ten_gv_val is not None else ''
                            if not ten_gv or ten_gv.lower() == 'none':
                                ten_gv = ma_gv

                            mail_gv_val = (
                                row.get('mail_gvcv') or row.get('email_gvcv') or row.get('mail gvcv') or
                                row.get('mail_gv') or row.get('email_gv') or row.get('mail gv') or
                                row.get('email_covan') or row.get('mail_covan') or
                                row.get('email_giangvien')
                            )
                            mail_gv = str(mail_gv_val).strip() if mail_gv_val is not None else ''
                            if mail_gv.lower() == 'none':
                                mail_gv = ''

                            covan, cv_created = User.objects.get_or_create(
                                username=ma_gv,
                                defaults={
                                    'full_name': ten_gv,
                                    'email': mail_gv,
                                    'role': 'covan',
                                }
                            )
                            if cv_created:
                                covan.set_password(ma_gv)
                                covan.save()
                            else:
                                updated = False
                                if ten_gv_val and covan.full_name != ten_gv:
                                    covan.full_name = ten_gv
                                    updated = True
                                if mail_gv_val and covan.email != mail_gv:
                                    covan.email = mail_gv
                                    updated = True
                                if covan.role != 'covan':
                                    covan.role = 'covan'
                                    updated = True
                                if updated:
                                    covan.save()

                    sv, _ = SinhVien.objects.update_or_create(
                        mssv=mssv,
                        defaults={
                            'ho_ten': ho_ten,
                            'email': email,
                            'khoa': khoa,
                            'nganh': nganh,
                            'trang_thai': trang_thai,
                            'user': sv_user,
                            'gioi_tinh': gioi_tinh,
                            'ngay_sinh': ngay_sinh,
                        }
                    )
                    # Gán lớp nếu có (tự động tạo lớp nếu chưa tồn tại)
                    ten_lop = str(row.get('lop', '') or '').strip()
                    if ten_lop and ten_lop.lower() != 'none':
                        from students.models import Lop
                        import re
                        m = re.search(r'\d{2}', ten_lop)
                        lop_khoa = '20' + m.group() if m else (khoa or '2022')
                        if lop_khoa.startswith('K') or lop_khoa.startswith('k'):
                            lop_khoa = lop_khoa[1:]

                        lop_obj, lop_created = Lop.objects.get_or_create(
                            ten_lop=ten_lop,
                            defaults={
                                'nganh': nganh,
                                'khoa': lop_khoa,
                            }
                        )

                        # Nếu lớp đã tồn tại nhưng chưa có ngành học thì gán ngành học
                        if not lop_created and not lop_obj.nganh:
                            lop_obj.nganh = nganh
                            lop_obj.save()

                        # Gán cố vấn cho lớp nếu cố vấn chưa nhận lớp nào khác
                        if covan:
                            already_advising = Lop.objects.filter(covan=covan).exclude(pk=lop_obj.pk).first()
                            if not already_advising:
                                lop_obj.covan = covan
                                lop_obj.save()

                        sv.lop = lop_obj
                        if covan:
                            sv.covan = covan
                        elif lop_obj.covan:
                            sv.covan = lop_obj.covan
                        sv.save()
                    elif covan:
                        sv.covan = covan
                        sv.save()
                    count += 1
                except Exception as e:
                    errors.append(f"MSSV {mssv}: {e}")

            messages.success(request, f'Import thành công {count} sinh viên.')
            if errors:
                messages.warning(request, f'{len(errors)} dòng lỗi: ' + '; '.join(errors[:5]))
            return redirect('students:sinhvien_list')
        except Exception as e:
            messages.error(request, f'Lỗi đọc file: {e}')
    return render(request, 'students/import.html', {'form': form, 'title': 'Import sinh viên', 'errors': errors})


@login_required
@role_required('giaovu', 'admin')
def import_monhoc(request):
    form = ImportCSVForm(request.POST or None, request.FILES or None)
    errors = []
    if request.method == 'POST' and form.is_valid():
        f = request.FILES['file']
        name = f.name.lower()
        count = 0
        try:
            if name.endswith('.csv'):
                decoded = f.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded))
                rows_data = list(reader)
            elif name.endswith(('.xlsx', '.xls')):
                wb = openpyxl.load_workbook(f)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                rows_data = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]
            else:
                messages.error(request, 'Định dạng file không hỗ trợ.')
                return render(request, 'students/import_monhoc.html', {'form': form, 'title': 'Import môn học', 'errors': errors})

            for row in rows_data:
                if not row:
                    continue
                ma_mh_val = row.get('ma_mh')
                if ma_mh_val is None:
                    continue
                ma_mh = str(ma_mh_val).strip()
                if not ma_mh or ma_mh.lower() == 'none':
                    continue
                try:
                    ten_mh_val = row.get('ten_mh')
                    ten_mh = str(ten_mh_val).strip() if ten_mh_val is not None else ''
                    if not ten_mh or ten_mh.lower() == 'none':
                        continue

                    so_tc_val = row.get('so_tc')
                    if so_tc_val is not None:
                        try:
                            so_tc = int(float(str(so_tc_val).strip()))
                        except ValueError:
                            so_tc = 3
                    else:
                        so_tc = 3

                    loai_val = row.get('loai')
                    loai = str(loai_val).strip() if loai_val is not None else 'bat_buoc'
                    if not loai or loai.lower() == 'none':
                        loai = 'bat_buoc'
                    
                    # Convert to standard choices value
                    loai_map = {
                        'bat_buoc': 'bat_buoc',
                        'bắt buộc': 'bat_buoc',
                        'tu_chon': 'tu_chon',
                        'tự chọn': 'tu_chon',
                        'dai_cuong': 'dai_cuong',
                        'đại cương': 'dai_cuong',
                        'chuyen_nganh': 'chuyen_nganh',
                        'chuyên ngành': 'chuyen_nganh',
                    }
                    loai = loai_map.get(loai.lower(), loai)

                    mo_ta_val = row.get('mo_ta')
                    mo_ta = str(mo_ta_val).strip() if mo_ta_val is not None else ''
                    if mo_ta.lower() == 'none':
                        mo_ta = ''

                    ma_nganh_val = row.get('ma_nganh') or row.get('ngành') or row.get('nganh') or row.get('chương trình') or row.get('chuong_trinh')
                    nganh = None
                    if ma_nganh_val:
                        ma_nganh = str(ma_nganh_val).strip()
                        if ma_nganh and ma_nganh.lower() != 'none':
                            nganh = Nganh.objects.filter(ma_nganh=ma_nganh).first()

                    nam_hoc_val = row.get('nam_hoc_ctdt') or row.get('năm học') or row.get('nam_hoc') or row.get('năm') or row.get('nam')
                    nam_hoc_ctdt = None
                    if nam_hoc_val is not None:
                        try:
                            nam_hoc_ctdt = int(float(str(nam_hoc_val).strip()))
                        except ValueError:
                            pass

                    hoc_ky_val = row.get('hoc_ky_ctdt') or row.get('học kỳ') or row.get('hoc_ky') or row.get('kỳ') or row.get('ky') or row.get('hk')
                    hoc_ky_ctdt = None
                    if hoc_ky_val is not None:
                        try:
                            hoc_ky_ctdt = int(float(str(hoc_ky_val).strip()))
                        except ValueError:
                            pass

                    MonHoc.objects.update_or_create(
                        ma_mh=ma_mh,
                        defaults={
                            'ten_mh': ten_mh,
                            'so_tc': so_tc,
                            'loai': loai,
                            'mo_ta': mo_ta,
                            'nganh': nganh,
                            'nam_hoc_ctdt': nam_hoc_ctdt,
                            'hoc_ky_ctdt': hoc_ky_ctdt,
                        }
                    )
                    count += 1
                except Exception as e:
                    errors.append(f"Mã MH {ma_mh}: {e}")

            messages.success(request, f'Import thành công {count} môn học.')
            if errors:
                messages.warning(request, f'{len(errors)} dòng lỗi: ' + '; '.join(errors[:5]))
            return redirect('students:monhoc_list')
        except Exception as e:
            messages.error(request, f'Lỗi đọc file: {e}')
    return render(request, 'students/import_monhoc.html', {'form': form, 'title': 'Import chương trình đào tạo', 'errors': errors})


# ---- Export Excel ----
@login_required
def export_sinhvien(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sinh viên'
    headers = ['MSSV', 'Họ tên', 'Email', 'Ngành', 'Khóa', 'Lớp', 'Trạng thái']
    ws.append(headers)
    for sv in SinhVien.objects.select_related('nganh').all():
        ws.append([sv.mssv, sv.ho_ten, sv.email,
                   sv.nganh.ten_nganh if sv.nganh else '',
                   sv.khoa, sv.lop, sv.get_trang_thai_display()])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sinh_vien.xlsx"'
    wb.save(response)
    return response
