import openpyxl
from io import BytesIO
from django.test import TestCase, Client
from django.urls import reverse
from accounts.models import CustomUser
from students.models import SinhVien, Lop, Nganh, HocKy, MonHoc
from results.models import KetQuaHocTap
from academic_warnings.models import CanhBaoHocVu
from dashboard.views import generate_excel_bytes

class ExcelExportTestCase(TestCase):
    def setUp(self):
        self.client = Client()
        
        # Create a giáo vụ user to run exports
        self.admin_user = CustomUser.objects.create_user(
            username='admin1', password='password', role='giaovu'
        )
        
        # Create Nganh
        self.nganh = Nganh.objects.create(ma_nganh='CNTT', ten_nganh='Cong nghe thong tin')
        
        # Create Lops
        self.lopA = Lop.objects.create(ten_lop='CNTK19A', nganh=self.nganh, khoa='K2019')
        self.lopB = Lop.objects.create(ten_lop='CNTK19B', nganh=self.nganh, khoa='K2019')
        
        # Create Students
        self.sv1 = SinhVien.objects.create(
            mssv='110122001', ho_ten='Student A', lop=self.lopA, nganh=self.nganh, khoa='K2019'
        )
        self.sv2 = SinhVien.objects.create(
            mssv='110122002', ho_ten='Student B', lop=self.lopB, nganh=self.nganh, khoa='K2019'
        )
        
        # Create Semesters
        self.hk1 = HocKy.objects.create(ky='1', nam_hoc='2023-2024')
        self.hk2 = HocKy.objects.create(ky='2', nam_hoc='2023-2024')
        
        # Create Course
        self.mh = MonHoc.objects.create(ma_mh='MH01', ten_mh='Web Programming', so_tc=3)
        
        # Add Results
        # Student A: active in HK1 & HK2
        KetQuaHocTap.objects.create(
            sinh_vien=self.sv1, mon_hoc=self.mh, hoc_ky=self.hk1,
            diem_qt=8.0, diem_thi=8.0, diem_tk=8.0, lan_hoc=1
        )
        KetQuaHocTap.objects.create(
            sinh_vien=self.sv1, mon_hoc=self.mh, hoc_ky=self.hk2,
            diem_qt=7.0, diem_thi=7.0, diem_tk=7.0, lan_hoc=1
        )
        
        # Student B: active in HK1 only
        KetQuaHocTap.objects.create(
            sinh_vien=self.sv2, mon_hoc=self.mh, hoc_ky=self.hk1,
            diem_qt=6.0, diem_thi=6.0, diem_tk=6.0, lan_hoc=1
        )

    def test_excel_export_unfiltered(self):
        # Unfiltered output should contain one row per student, without "TC tích lũy", grouped by class
        hks = HocKy.objects.all()
        svs = SinhVien.objects.all()
        
        excel_data = generate_excel_bytes(svs, hks, nam_hoc_filter='', ky_filter='')
        wb = openpyxl.load_workbook(BytesIO(excel_data))
        ws = wb.active
        
        # Verify title and filters in header
        self.assertEqual(ws['A1'].value, 'BÁO CÁO TỔNG HỢP KẾT QUẢ HỌC TẬP & CẢNH BÁO HỌC VỤ')
        self.assertEqual(ws['A2'].value, 'Tất cả')
        
        # Find headers row
        header_row = None
        for row in range(4, 10):
            if ws.cell(row=row, column=1).value == 'MSSV':
                header_row = row
                break
        self.assertIsNotNone(header_row)
        
        # Check header values
        headers = [ws.cell(row=header_row, column=col).value for col in range(1, 9)]
        expected_headers = ['MSSV', 'Họ tên', 'Lớp', 'Ngành', 'ĐTBCTL (Hệ 10)', 'ĐTBCTL (Hệ 4)', 'Trạng thái hiện tại', 'Cảnh báo học vụ mới nhất']
        self.assertEqual(headers, expected_headers)
        
        # Cumulative credits 'TC tích lũy' must NOT be in headers
        all_header_values = [ws.cell(row=header_row, column=col).value for col in range(1, 20) if ws.cell(row=header_row, column=col).value]
        self.assertNotIn('TC tích lũy', all_header_values)
        
        # Class headers and students order checking
        # Group CNTK19A (sv1) comes first, then CNTK19B (sv2)
        # Row layout:
        # Lớp: CNTK19A
        # sv1 row
        # (empty row separator)
        # Lớp: CNTK19B
        # sv2 row
        
        class_a_row = header_row + 1
        self.assertEqual(ws.cell(row=class_a_row, column=1).value, 'Lớp: CNTK19A')
        
        sv1_row = class_a_row + 1
        self.assertEqual(ws.cell(row=sv1_row, column=1).value, self.sv1.mssv)
        self.assertEqual(ws.cell(row=sv1_row, column=2).value, self.sv1.ho_ten)
        
        sep_row = sv1_row + 1
        self.assertEqual(ws.cell(row=sep_row, column=1).value, None)
        
        class_b_row = sep_row + 1
        self.assertEqual(ws.cell(row=class_b_row, column=1).value, 'Lớp: CNTK19B')
        
        sv2_row = class_b_row + 1
        self.assertEqual(ws.cell(row=sv2_row, column=1).value, self.sv2.mssv)

    def test_excel_export_filtered_by_year(self):
        # Filtered output should contain separate columns for each semester, without "TC tích lũy" and on a single row per student
        hks = HocKy.objects.filter(nam_hoc='2023-2024')
        svs = SinhVien.objects.all()
        
        excel_data = generate_excel_bytes(svs, hks, nam_hoc_filter='2023-2024', ky_filter='')
        wb = openpyxl.load_workbook(BytesIO(excel_data))
        ws = wb.active
        
        # Find headers row
        header_row = None
        for row in range(4, 10):
            if ws.cell(row=row, column=1).value == 'MSSV':
                header_row = row
                break
        self.assertIsNotNone(header_row)
        
        # Check header values
        headers = [ws.cell(row=header_row, column=col).value for col in range(1, 15)]
        expected_headers = [
            'MSSV', 'Họ tên', 'Lớp', 'Ngành',
            'ĐTBCHK HK1 (Hệ 10)', 'ĐTBCHK HK1 (Hệ 4)', 'ĐTBCTL (Hệ 10)', 'ĐTBCTL (Hệ 4)', 'Cảnh báo HK1',
            'ĐTBCHK HK2 (Hệ 10)', 'ĐTBCHK HK2 (Hệ 4)', 'ĐTBCTL (Hệ 10)', 'ĐTBCTL (Hệ 4)', 'Cảnh báo HK2'
        ]
        self.assertEqual(headers, expected_headers)
        
        # 'TC tích lũy' must NOT be in headers
        all_header_values = [ws.cell(row=header_row, column=col).value for col in range(1, 30) if ws.cell(row=header_row, column=col).value]
        self.assertNotIn('TC tích lũy', all_header_values)
        
        # Class CNTK19A is first
        class_a_row = header_row + 1
        self.assertEqual(ws.cell(row=class_a_row, column=1).value, 'Lớp: CNTK19A')
        
        # sv1 is on class_a_row + 1. Single row contains HK1 and HK2 GPAs.
        sv1_row = class_a_row + 1
        self.assertEqual(ws.cell(row=sv1_row, column=1).value, self.sv1.mssv)
        # HK1 values (active)
        self.assertEqual(ws.cell(row=sv1_row, column=5).value, 8.00) # ĐTBCHK HK1 10
        self.assertEqual(ws.cell(row=sv1_row, column=6).value, 3.50) # ĐTBCHK HK1 4
        # HK2 values (active)
        self.assertEqual(ws.cell(row=sv1_row, column=10).value, 7.00) # ĐTBCHK HK2 10
        self.assertEqual(ws.cell(row=sv1_row, column=11).value, 3.00) # ĐTBCHK HK2 4
        
        # Separator row
        sep_row = sv1_row + 1
        self.assertEqual(ws.cell(row=sep_row, column=1).value, None)
        
        # Class CNTK19B is next
        class_b_row = sep_row + 1
        self.assertEqual(ws.cell(row=class_b_row, column=1).value, 'Lớp: CNTK19B')
        
        # sv2 is on class_b_row + 1. Only active in HK1.
        sv2_row = class_b_row + 1
        self.assertEqual(ws.cell(row=sv2_row, column=1).value, self.sv2.mssv)
        # HK1 values (active)
        self.assertEqual(ws.cell(row=sv2_row, column=5).value, 6.00) # ĐTBCHK HK1 10
        self.assertEqual(ws.cell(row=sv2_row, column=6).value, 2.00) # ĐTBCHK HK1 4
        # HK2 values (not active) -> should display '-' for semester GPA
        self.assertEqual(ws.cell(row=sv2_row, column=10).value, '-') # ĐTBCHK HK2 10
        self.assertEqual(ws.cell(row=sv2_row, column=11).value, '-') # ĐTBCHK HK2 4

    def test_view_export_bao_cao_excel(self):
        self.client.login(username='admin1', password='password')
        response = self.client.get(reverse('dashboard:export_bao_cao'), {'nam_hoc': '2023-2024'})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertTrue(len(response.content) > 0)
