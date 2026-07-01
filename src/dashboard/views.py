from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, Avg, Q
from io import BytesIO
import json

from students.models import SinhVien, HocKy, MonHoc, Nganh
from results.models import KetQuaHocTap
from results.utils import (tinh_dtbctl, tinh_dtbchk,
                            thong_ke_hocky, get_phan_phoi_diem, tinh_canh_bao_som)
from academic_warnings.models import CanhBaoHocVu


@login_required
def index(request):
    """Điều hướng đến dashboard phù hợp theo vai trò."""
    role = request.user.role
    if role == 'sinhvien':
        return dashboard_sinhvien(request)
    elif role == 'covan':
        return dashboard_covan(request)
    elif role in ('giaovu', 'admin'):
        return dashboard_giaovu(request)
    return render(request, 'dashboard/index.html')


@login_required
def dashboard_sinhvien(request):
    """Dashboard dành cho sinh viên."""
    sv = None
    try:
        sv = request.user.sinh_vien
    except Exception:
        pass

    # Thử tìm qua SinhVien.objects nếu reverse relation chưa cache
    if sv is None:
        from students.models import SinhVien as SV
        sv = SV.objects.filter(user=request.user).first()

    if sv is None:
        return render(request, 'dashboard/no_profile.html')

    hoc_ky_hien_tai = HocKy.objects.filter(la_hien_tai=True).first()
    dtbctl, dtbctl_4, tc_tl, tc_da_hoc = tinh_dtbctl(sv)

    # ĐTBCHK từng học kỳ
    hockys = HocKy.objects.filter(ket_qua__sinh_vien=sv).distinct().order_by('nam_hoc', 'ky')
    gpa_labels = []
    gpa_data = []
    for hk in hockys:
        dtbchk_10, dtbchk_4, tc, _ = tinh_dtbchk(sv, hk)
        if tc > 0:
            gpa_labels.append(str(hk))
            gpa_data.append(dtbchk_10)

    # Thống kê học kỳ hiện tại
    thong_ke = None
    if hoc_ky_hien_tai:
        thong_ke = thong_ke_hocky(sv, hoc_ky_hien_tai)

    # Phân phối điểm
    phan_phoi = get_phan_phoi_diem(sinh_vien=sv)

    # Cảnh báo
    canh_baos = CanhBaoHocVu.objects.filter(sinh_vien=sv).order_by('-ngay_tao')[:5]

    # Kết quả gần nhất
    ket_qua_gan = KetQuaHocTap.objects.filter(sinh_vien=sv).select_related(
        'mon_hoc', 'hoc_ky').order_by('-hoc_ky__nam_hoc', '-hoc_ky__ky')[:10]

    context = {
        'sv': sv,
        'dtbctl': dtbctl,
        'dtbctl_4': dtbctl_4,
        'tc_tl': tc_tl,
        'tc_da_hoc': tc_da_hoc,
        'hoc_ky_hien_tai': hoc_ky_hien_tai,
        'thong_ke': thong_ke,
        'canh_baos': canh_baos,
        'canh_bao_som': tinh_canh_bao_som(sv),
        'ket_qua_gan': ket_qua_gan,
        'gpa_labels': json.dumps(gpa_labels),
        'gpa_data': json.dumps(gpa_data),
        'phan_phoi_labels': json.dumps(list(phan_phoi.keys())),
        'phan_phoi_data': json.dumps(list(phan_phoi.values())),
    }
    return render(request, 'dashboard/dashboard_sv.html', context)


@login_required
def dashboard_covan(request):
    """Dashboard dành cho cố vấn học tập."""
    svs = SinhVien.objects.filter(lop__covan=request.user)
    total_sv = svs.count()
    # sv_canh_bao: số sinh viên đang có cảnh báo học vụ (lần 1 hoặc lần 2)
    sv_canh_bao = CanhBaoHocVu.objects.filter(
        sinh_vien__lop__covan=request.user, muc_canh_bao='canh_bao'
    ).values('sinh_vien').distinct().count()
    # sv_dang_hoc: bao gồm cả dang_hoc và canh_bao (tương thích dữ liệu cũ)
    sv_dang_hoc = svs.filter(trang_thai__in=['dang_hoc', 'canh_bao']).count()

    # Cảnh báo mới của kỳ gần nhất
    latest_hk = HocKy.objects.filter(ket_qua__diem_tk__isnull=False).distinct().order_by('-nam_hoc', '-ky').first()
    if not latest_hk:
        latest_hk = HocKy.objects.order_by('-nam_hoc', '-ky').first()

    canh_baos_moi = CanhBaoHocVu.objects.filter(
        sinh_vien__lop__covan=request.user
    )
    if latest_hk:
        canh_baos_moi = canh_baos_moi.filter(hoc_ky=latest_hk)
    canh_baos_moi = canh_baos_moi.exclude(nguoi_dung_an=request.user).select_related('sinh_vien', 'hoc_ky').order_by('-ngay_tao')[:10]

    # Thống kê theo mức cảnh báo
    cb_stats = CanhBaoHocVu.objects.filter(sinh_vien__lop__covan=request.user).values(
        'muc_canh_bao').annotate(count=Count('id'))
    cb_labels = [d['muc_canh_bao'] for d in cb_stats]
    cb_data = [d['count'] for d in cb_stats]

    # Danh sách học kỳ để cố vấn lựa chọn lọc GPA
    hockys = HocKy.objects.all().order_by('-nam_hoc', '-ky')
    hk_id = request.GET.get('hoc_ky', '')

    if hk_id:
        selected_hk = HocKy.objects.filter(pk=hk_id).first()
    else:
        selected_hk = HocKy.objects.filter(la_hien_tai=True).first()
        if not selected_hk and hockys.exists():
            selected_hk = hockys[0]

    sv_gpa_thap = []
    if selected_hk:
        for sv in svs:
            dtbchk_10, dtbchk_4, tc, _ = tinh_dtbchk(sv, selected_hk)
            if tc > 0 and dtbchk_4 < 2.0:
                sv_gpa_thap.append({'sv': sv, 'gpa': dtbchk_4, 'tc': tc})
        sv_gpa_thap.sort(key=lambda x: x['gpa'])

    # Phân tích cảnh báo sớm cho các sinh viên lớp cố vấn phụ trách
    student_list = list(svs)
    student_ids = [sv.id for sv in student_list]

    # Query 1: Fetch all KetQuaHocTap with select_related for all students in student_ids
    results_list = list(
        KetQuaHocTap.objects.filter(sinh_vien_id__in=student_ids, diem_tk__isnull=False)
        .select_related('mon_hoc', 'hoc_ky')
        .order_by('hoc_ky__nam_hoc', 'hoc_ky__ky')
    )
    prefetch_results = {}
    for r in results_list:
        if r.sinh_vien_id not in prefetch_results:
            prefetch_results[r.sinh_vien_id] = []
        prefetch_results[r.sinh_vien_id].append(r)

    # Query 2: Fetch all CanhBaoHocVu for these students
    warnings_list = list(CanhBaoHocVu.objects.filter(sinh_vien_id__in=student_ids))
    prefetch_warnings = {}
    for cb in warnings_list:
        prefetch_warnings[(cb.sinh_vien_id, cb.hoc_ky_id)] = cb

    cbs_stats = {'safe': 0, 'monitor': 0, 'warning_1': 0, 'warning_2': 0}
    sv_risk_list = []
    for sv in student_list:
        analysis = tinh_canh_bao_som(sv, prefetch_results=prefetch_results, prefetch_warnings=prefetch_warnings)
        lvl = analysis['muc_nguy_co']
        if lvl in cbs_stats:
            cbs_stats[lvl] += 1
        if lvl in ['warning_1', 'warning_2', 'monitor']:
            sv_risk_list.append(analysis)
            
    level_order = {'warning_2': 0, 'warning_1': 1, 'monitor': 2, 'safe': 3}
    sv_risk_list.sort(key=lambda x: level_order.get(x['muc_nguy_co'], 4))

    # Phân bố trạng thái sinh viên
    trang_thai_stats = svs.values('trang_thai').annotate(count=Count('id'))

    context = {
        'total_sv': total_sv,
        'sv_canh_bao': sv_canh_bao,
        'sv_dang_hoc': sv_dang_hoc,
        'canh_baos_moi': canh_baos_moi,
        'sv_gpa_thap': sv_gpa_thap[:10],
        'sv_risk_list': sv_risk_list[:10],
        'cbs_stats': cbs_stats,
        'hockys': hockys,
        'selected_hk': selected_hk,
        'cb_labels': json.dumps([d['muc_canh_bao'] for d in cb_stats]),
        'cb_data': json.dumps([d['count'] for d in cb_stats]),
        'tt_labels': json.dumps([d['trang_thai'] for d in trang_thai_stats]),
        'tt_data': json.dumps([d['count'] for d in trang_thai_stats]),
    }
    return render(request, 'dashboard/dashboard_covan.html', context)


@login_required
def dashboard_giaovu(request):
    """Dashboard dành cho giáo vụ và admin."""
    total_sv = SinhVien.objects.count()
    # sv_canh_bao: số sinh viên đang có cảnh báo học vụ (lần 1 hoặc lần 2) thuộc HK1 2025-2026
    sv_canh_bao = CanhBaoHocVu.objects.filter(
        muc_canh_bao='canh_bao',
        hoc_ky__ky='1',
        hoc_ky__nam_hoc='2025-2026'
    ).values('sinh_vien').distinct().count()
    # sv_dang_hoc: bao gồm cả dang_hoc và canh_bao (tương thích dữ liệu cũ)
    sv_dang_hoc = SinhVien.objects.filter(trang_thai__in=['dang_hoc', 'canh_bao']).count()
    total_mon = MonHoc.objects.count()
    total_canh_bao = CanhBaoHocVu.objects.filter(
        trang_thai='chua_xu_ly',
        muc_canh_bao='canh_bao',
        hoc_ky__ky='1',
        hoc_ky__nam_hoc='2025-2026'
    ).exclude(nguoi_dung_an=request.user).count()

    # Thống kê cảnh báo theo học kỳ
    hockys = HocKy.objects.order_by('nam_hoc', 'ky')[:8]
    cb_hk_labels = [str(hk) for hk in hockys]
    cb_hk_data = [CanhBaoHocVu.objects.filter(hoc_ky=hk).count() for hk in hockys]

    # Phân bố sinh viên theo ngành
    nganh_stats = SinhVien.objects.values('nganh__ten_nganh').annotate(count=Count('id')).order_by('-count')[:8]

    # Phân bố trạng thái
    tt_stats = SinhVien.objects.values('trang_thai').annotate(count=Count('id'))

    # Cảnh báo mới nhất (chỉ hiển thị HK1 2025-2026)
    canh_baos_moi = CanhBaoHocVu.objects.filter(
        trang_thai='chua_xu_ly',
        hoc_ky__ky='1',
        hoc_ky__nam_hoc='2025-2026'
    ).exclude(nguoi_dung_an=request.user).select_related(
        'sinh_vien', 'hoc_ky'
    ).order_by('-ngay_tao')[:10]

    # GPA trung bình theo học kỳ (tính từ dữ liệu)
    hoc_ky_hien_tai = HocKy.objects.filter(la_hien_tai=True).first()
    phan_phoi = get_phan_phoi_diem(hoc_ky=hoc_ky_hien_tai) if hoc_ky_hien_tai else {}

    context = {
        'total_sv': total_sv,
        'sv_canh_bao': sv_canh_bao,
        'sv_dang_hoc': sv_dang_hoc,
        'total_mon': total_mon,
        'total_canh_bao': total_canh_bao,
        'canh_baos_moi': canh_baos_moi,
        'hoc_ky_hien_tai': hoc_ky_hien_tai,
        'cb_hk_labels': json.dumps(cb_hk_labels),
        'cb_hk_data': json.dumps(cb_hk_data),
        'nganh_labels': json.dumps([d['nganh__ten_nganh'] or 'Chưa xác định' for d in nganh_stats]),
        'nganh_data': json.dumps([d['count'] for d in nganh_stats]),
        'tt_labels': json.dumps([d['trang_thai'] for d in tt_stats]),
        'tt_data': json.dumps([d['count'] for d in tt_stats]),
        'phan_phoi_labels': json.dumps(list(phan_phoi.keys())),
        'phan_phoi_data': json.dumps(list(phan_phoi.values())),
    }
    return render(request, 'dashboard/dashboard_giaovu.html', context)


@login_required
def bao_cao(request):
    """Trang báo cáo & thống kê."""
    nam_hocs = HocKy.objects.values_list('nam_hoc', flat=True).distinct().order_by('-nam_hoc')
    nam_hoc_filter = request.GET.get('nam_hoc', '')
    ky_filter = request.GET.get('ky', '')

    qs_sv = SinhVien.objects.all()
    qs_kq = KetQuaHocTap.objects.select_related('sinh_vien', 'mon_hoc', 'hoc_ky')
    qs_cb = CanhBaoHocVu.objects.select_related('sinh_vien', 'hoc_ky')

    if request.user.is_covan:
        qs_sv = qs_sv.filter(lop__covan=request.user)
        qs_kq = qs_kq.filter(sinh_vien__lop__covan=request.user)
        qs_cb = qs_cb.filter(sinh_vien__lop__covan=request.user)

    if nam_hoc_filter:
        qs_kq = qs_kq.filter(hoc_ky__nam_hoc=nam_hoc_filter)
        qs_cb = qs_cb.filter(hoc_ky__nam_hoc=nam_hoc_filter)

    if ky_filter:
        qs_kq = qs_kq.filter(hoc_ky__ky=ky_filter)
        qs_cb = qs_cb.filter(hoc_ky__ky=ky_filter)

    # Thống kê tổng hợp
    stats = {
        'total_sv': qs_sv.count(),
        # sv_dang_hoc: bao gồm cả dang_hoc và canh_bao (tương thích dữ liệu cũ)
        'sv_dang_hoc': qs_sv.filter(trang_thai__in=['dang_hoc', 'canh_bao']).count(),
        'sv_canh_bao': qs_cb.filter(muc_canh_bao='canh_bao').count(),
        'sv_canh_bao_2': qs_cb.filter(muc_canh_bao='canh_bao', so_lan_canh_bao=2).count(),
        'sv_canh_bao_3': qs_cb.filter(muc_canh_bao='buoc_thoi_hoc').count(),
        'total_kq': qs_kq.count(),
        'kq_dat': sum(1 for kq in qs_kq if kq.diem_tk is not None and kq.diem_tk >= 4.0),
    }

    # Biểu đồ: Cảnh báo theo lớp
    cb_lop_stats = qs_cb.values('sinh_vien__lop__ten_lop').annotate(count=Count('id')).order_by('-count')[:15]
    cb_lop_labels = [item['sinh_vien__lop__ten_lop'] or 'Chưa xếp lớp' for item in cb_lop_stats]
    cb_lop_data = [item['count'] for item in cb_lop_stats]

    # Biểu đồ: Cảnh báo theo ngành
    cb_nganh_stats = qs_cb.values('sinh_vien__nganh__ten_nganh').annotate(count=Count('id')).order_by('-count')[:15]
    cb_nganh_labels = [item['sinh_vien__nganh__ten_nganh'] or 'Chưa xếp ngành' for item in cb_nganh_stats]
    cb_nganh_data = [item['count'] for item in cb_nganh_stats]

    # Biểu đồ: Cảnh báo theo khoa
    cb_khoa_stats = qs_cb.values('sinh_vien__nganh__khoa').annotate(count=Count('id')).order_by('-count')[:15]
    cb_khoa_labels = [item['sinh_vien__nganh__khoa'] or 'Chưa xác định khoa' for item in cb_khoa_stats]
    cb_khoa_data = [item['count'] for item in cb_khoa_stats]

    # Biểu đồ: Cảnh báo theo khóa (cohort)
    cb_khoa_hoc_stats = qs_cb.values('sinh_vien__khoa').annotate(count=Count('id')).order_by('-count')[:15]
    cb_khoa_hoc_labels = [f"Khóa {item['sinh_vien__khoa']}" if item['sinh_vien__khoa'] else 'Chưa xác định' for item in cb_khoa_hoc_stats]
    cb_khoa_hoc_data = [item['count'] for item in cb_khoa_hoc_stats]

    # Biểu đồ: Sinh viên theo lớp
    sv_lop_stats = qs_sv.values('lop__ten_lop').annotate(count=Count('id')).order_by('-count')[:15]
    sv_lop_labels = [item['lop__ten_lop'] or 'Chưa xếp lớp' for item in sv_lop_stats]
    sv_lop_data = [item['count'] for item in sv_lop_stats]

    # Biểu đồ: Sinh viên theo ngành
    sv_nganh_stats = qs_sv.values('nganh__ten_nganh').annotate(count=Count('id')).order_by('-count')[:15]
    sv_nganh_labels = [item['nganh__ten_nganh'] or 'Chưa xếp ngành' for item in sv_nganh_stats]
    sv_nganh_data = [item['count'] for item in sv_nganh_stats]

    # Biểu đồ: Sinh viên theo khoa
    sv_khoa_stats = qs_sv.values('nganh__khoa').annotate(count=Count('id')).order_by('-count')[:15]
    sv_khoa_labels = [item['nganh__khoa'] or 'Chưa xác định khoa' for item in sv_khoa_stats]
    sv_khoa_data = [item['count'] for item in sv_khoa_stats]

    # Biểu đồ: Sinh viên theo khóa (cohort)
    sv_khoa_hoc_stats = qs_sv.values('khoa').annotate(count=Count('id')).order_by('-count')[:15]
    sv_khoa_hoc_labels = [f"Khóa {item['khoa']}" if item['khoa'] else 'Chưa xác định' for item in sv_khoa_hoc_stats]
    sv_khoa_hoc_data = [item['count'] for item in sv_khoa_hoc_stats]

    # Biểu đồ: Trạng thái sinh viên
    tt_stats = qs_sv.values('trang_thai').annotate(count=Count('id'))

    context = {
        'nam_hocs': nam_hocs,
        'nam_hoc_filter': nam_hoc_filter,
        'ky_filter': ky_filter,
        'stats': stats,
        'cb_lop_labels': json.dumps(cb_lop_labels),
        'cb_lop_data': json.dumps(cb_lop_data),
        'cb_nganh_labels': json.dumps(cb_nganh_labels),
        'cb_nganh_data': json.dumps(cb_nganh_data),
        'cb_khoa_labels': json.dumps(cb_khoa_labels),
        'cb_khoa_data': json.dumps(cb_khoa_data),
        'cb_khoa_hoc_labels': json.dumps(cb_khoa_hoc_labels),
        'cb_khoa_hoc_data': json.dumps(cb_khoa_hoc_data),
        'sv_lop_labels': json.dumps(sv_lop_labels),
        'sv_lop_data': json.dumps(sv_lop_data),
        'sv_nganh_labels': json.dumps(sv_nganh_labels),
        'sv_nganh_data': json.dumps(sv_nganh_data),
        'sv_khoa_labels': json.dumps(sv_khoa_labels),
        'sv_khoa_data': json.dumps(sv_khoa_data),
        'sv_khoa_hoc_labels': json.dumps(sv_khoa_hoc_labels),
        'sv_khoa_hoc_data': json.dumps(sv_khoa_hoc_data),
        'tt_labels': json.dumps([d['trang_thai'] for d in tt_stats]),
        'tt_data': json.dumps([d['count'] for d in tt_stats]),
    }
    return render(request, 'dashboard/bao_cao.html', context)


def generate_excel_bytes(svs_queryset, hks, nam_hoc_filter, ky_filter, recipient_name=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from results.utils import tinh_dtbctl, tinh_dtbchk

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Báo cáo tổng hợp'

    # Stylings
    title_font = Font(name='Segoe UI', size=16, bold=True, color='1F4E78')
    subtitle_font = Font(name='Segoe UI', size=11, italic=True, color='566573')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # Slate/Navy Blue
    
    class_header_font = Font(name='Segoe UI', size=11, bold=True, color='1F4E78')
    class_header_fill = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid') # Soft blue highlight
    
    regular_font = Font(name='Segoe UI', size=11)
    
    thin_border_side = Side(border_style='thin', color='D5D8DC')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    # Title info
    ws.append(['BÁO CÁO TỔNG HỢP KẾT QUẢ HỌC TẬP & CẢNH BÁO HỌC VỤ'])
    ws.cell(row=1, column=1).font = title_font
    
    filter_desc = []
    if nam_hoc_filter:
        filter_desc.append(f"Năm học: {nam_hoc_filter}")
    if ky_filter:
        filter_desc.append(f"Học kỳ: {ky_filter}")
    ws.append([f"{', '.join(filter_desc) if filter_desc else 'Tất cả'}"])
    ws.cell(row=2, column=1).font = subtitle_font
    
    row_offset = 3
    if recipient_name:
        ws.append([f"Người nhận: {recipient_name}"])
        ws.cell(row=3, column=1).font = subtitle_font
        row_offset = 4
    
    ws.append([]) # Empty row

    # Sort queryset by class name and mssv
    svs_queryset = svs_queryset.order_by('lop__ten_lop', 'mssv')

    is_filtered = bool(nam_hoc_filter or ky_filter)
    hks_ordered = hks.order_by('nam_hoc', 'ky')

    if is_filtered:
        headers = ['MSSV', 'Họ tên', 'Lớp', 'Ngành']
        for hk in hks_ordered:
            headers.extend([
                f'ĐTBCHK HK{hk.ky} (Hệ 10)',
                f'ĐTBCHK HK{hk.ky} (Hệ 4)',
                f'ĐTBCTL (Hệ 10)',
                f'ĐTBCTL (Hệ 4)',
                f'Cảnh báo HK{hk.ky}'
            ])
    else:
        headers = ['MSSV', 'Họ tên', 'Lớp', 'Ngành', 'ĐTBCTL (Hệ 10)', 'ĐTBCTL (Hệ 4)', 'Trạng thái hiện tại', 'Cảnh báo học vụ mới nhất']

    ws.append(headers)
    header_row_idx = ws.max_row
    
    # Format headers
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    current_lop = None

    for sv in svs_queryset:
        lop_name = sv.lop.ten_lop if sv.lop else 'Chưa xếp lớp'
        
        # Check for class change to append separator
        if lop_name != current_lop:
            if current_lop is not None:
                ws.append([]) # Empty row separator
            
            # Append class header row
            ws.append([f"Lớp: {lop_name}"])
            class_row_idx = ws.max_row
            # Merge class header row across all columns
            ws.merge_cells(start_row=class_row_idx, start_column=1, end_row=class_row_idx, end_column=len(headers))
            # Format merged class header cell
            class_cell = ws.cell(row=class_row_idx, column=1)
            class_cell.font = class_header_font
            class_cell.fill = class_header_fill
            class_cell.alignment = align_left
            
            # Apply border to all cells in the merged row to make it neat
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=class_row_idx, column=col_idx).border = thin_border
                
            current_lop = lop_name

        if is_filtered:
            row_data = [sv.mssv, sv.ho_ten, lop_name, sv.nganh.ten_nganh if sv.nganh else '']
            for hk in hks_ordered:
                dtbchk, dtbchk_4, tc_hk, _ = tinh_dtbchk(sv, hk)
                dtbctl, dtbctl_4, tc_ctl_da_hoc, _ = tinh_dtbctl(sv, hk)
                cb = CanhBaoHocVu.objects.filter(sinh_vien=sv, hoc_ky=hk).first()
                cb_str = cb.get_muc_canh_bao_display() if cb else '-'
                
                # Clean value formatting: display '-' if student was not active
                val_dtbchk = dtbchk if tc_hk > 0 else '-'
                val_dtbchk_4 = dtbchk_4 if tc_hk > 0 else '-'
                val_dtbctl = dtbctl if tc_ctl_da_hoc > 0 else '-'
                val_dtbctl_4 = dtbctl_4 if tc_ctl_da_hoc > 0 else '-'
                
                row_data.extend([val_dtbchk, val_dtbchk_4, val_dtbctl, val_dtbctl_4, cb_str])
                
            ws.append(row_data)
            data_row_idx = ws.max_row
            # Format data row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=data_row_idx, column=col_idx)
                cell.font = regular_font
                cell.border = thin_border
                # Alignments
                if col_idx in [1, 3]: # MSSV, Lớp
                    cell.alignment = align_center
                elif col_idx in [2, 4]: # Họ tên, Ngành
                    cell.alignment = align_left
                else:
                    rel_idx = (col_idx - 5) % 5
                    if rel_idx == 4: # Cảnh báo
                        cell.alignment = align_center
                    else: # GPAs
                        cell.alignment = align_right
        else:
            dtbctl, dtbctl_4, _, _ = tinh_dtbctl(sv)
            cb = CanhBaoHocVu.objects.filter(sinh_vien=sv).order_by('-ngay_tao').first()
            cb_str = cb.get_muc_canh_bao_display() if cb else '-'
            row_data = [
                sv.mssv, sv.ho_ten, lop_name, sv.nganh.ten_nganh if sv.nganh else '',
                dtbctl, dtbctl_4, sv.get_trang_thai_display(), cb_str
            ]
            ws.append(row_data)
            data_row_idx = ws.max_row
            # Format data row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=data_row_idx, column=col_idx)
                cell.font = regular_font
                cell.border = thin_border
                if col_idx in [1, 3, 7, 8]: # MSSV, Lớp, Trạng thái, Cảnh báo
                    cell.alignment = align_center
                elif col_idx in [5, 6]: # Numeric GPAs
                    cell.alignment = align_right
                else: # Họ tên, Ngành
                    cell.alignment = align_left

    # Auto-adjust column widths, ignoring the title and header/separator rows
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Skip title rows and class separator rows for width calculation
            val = str(cell.value or '')
            if cell.row <= (row_offset + 2) or val.startswith("Lớp: "):
                continue
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@login_required
def export_bao_cao_excel(request):
    """Xuất báo cáo tổng hợp ra Excel."""
    nam_hoc_filter = request.GET.get('nam_hoc', '')
    ky_filter = request.GET.get('ky', '')

    hks = HocKy.objects.all()
    if nam_hoc_filter:
        hks = hks.filter(nam_hoc=nam_hoc_filter)
    if ky_filter:
        hks = hks.filter(ky=ky_filter)

    svs = SinhVien.objects.select_related('nganh', 'lop').all()
    if request.user.is_covan:
        svs = svs.filter(covan=request.user)

    if nam_hoc_filter or ky_filter:
        # Lọc những sinh viên có kết quả học tập hoặc cảnh báo trong các học kỳ này
        sv_ids = set()
        q_kq = KetQuaHocTap.objects.filter(hoc_ky__in=hks)
        sv_ids.update(q_kq.values_list('sinh_vien_id', flat=True))
        q_cb = CanhBaoHocVu.objects.filter(hoc_ky__in=hks)
        sv_ids.update(q_cb.values_list('sinh_vien_id', flat=True))
        svs = svs.filter(id__in=sv_ids)

    excel_data = generate_excel_bytes(svs, hks, nam_hoc_filter, ky_filter)
    
    response = HttpResponse(excel_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="bao_cao_hoc_tap.xlsx"'
    return response


@login_required
def gui_bao_cao_covan(request):
    """Gửi báo cáo Excel lớp quản lý cho Cố vấn học tập tương ứng."""
    if request.method != 'POST':
        return HttpResponse("Method not allowed", status=405)

    if not (request.user.is_giaovu or request.user.is_admin):
        from django.contrib import messages
        messages.error(request, 'Bạn không có quyền thực hiện hành động này.')
        return redirect('dashboard:bao_cao')

    nam_hoc_filter = request.GET.get('nam_hoc', '')
    ky_filter = request.GET.get('ky', '')

    hks = HocKy.objects.all()
    if nam_hoc_filter:
        hks = hks.filter(nam_hoc=nam_hoc_filter)
    if ky_filter:
        hks = hks.filter(ky=ky_filter)

    from accounts.models import CustomUser
    from django.core.mail import EmailMessage
    from django.conf import settings
    from django.contrib import messages

    covans = CustomUser.objects.filter(role='covan', email__isnull=False).exclude(email='')

    sent_count = 0
    errors = []

    for covan in covans:
        # Lấy tất cả sinh viên thuộc lớp do cố vấn này quản lý
        svs = SinhVien.objects.filter(lop__covan=covan).select_related('nganh', 'lop')
        if not svs.exists():
            continue

        if nam_hoc_filter or ky_filter:
            sv_ids = set()
            q_kq = KetQuaHocTap.objects.filter(hoc_ky__in=hks)
            sv_ids.update(q_kq.values_list('sinh_vien_id', flat=True))
            q_cb = CanhBaoHocVu.objects.filter(hoc_ky__in=hks)
            sv_ids.update(q_cb.values_list('sinh_vien_id', flat=True))
            svs = svs.filter(id__in=sv_ids)

        if not svs.exists():
            continue

        # Tạo Excel cho riêng cố vấn này
        excel_data = generate_excel_bytes(
            svs, hks, nam_hoc_filter, ky_filter, 
            recipient_name=f"Cố vấn học tập {covan.full_name}"
        )

        if nam_hoc_filter and ky_filter:
            filter_text = f"Học kỳ {ky_filter} năm học {nam_hoc_filter}"
        elif nam_hoc_filter:
            filter_text = f"Năm học {nam_hoc_filter}"
        elif ky_filter:
            filter_text = f"Học kỳ {ky_filter}"
        else:
            filter_text = f"Tất cả học kỳ"

        subject = f"[TVU] Báo cáo học tập & Cảnh báo học vụ lớp phụ trách - {filter_text}"
        body = f"""Kính gửi Thầy/Cô cố vấn học tập {covan.full_name},

Hệ thống Quản lý Học vụ Trường Kỹ thuật và Công nghệ - Đại học Trà Vinh gửi kèm báo cáo kết quả học tập và cảnh báo học vụ của lớp/sinh viên do Thầy/Cô phụ trách.

- Kỳ/năm báo cáo: {filter_text}
- File đính kèm: Báo cáo định dạng Excel (.xlsx).

Kính đề nghị Thầy/Cô tải file đính kèm để xem chi tiết và thực hiện tư vấn học vụ kịp thời cho các sinh viên.

Trân trọng,
Giáo vụ Trường Kỹ thuật và Công nghệ
"""
        try:
            email = EmailMessage(
                subject=subject,
                body=body,
                from_email=settings.DEFAULT_FROM_EMAIL or 'giaovu@tvu.edu.vn',
                to=[covan.email],
            )
            email.attach(f"bao_cao_hoc_tap_{covan.username}.xlsx", excel_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            email.send()
            sent_count += 1
        except Exception as e:
            errors.append(f"{covan.full_name}: {str(e)}")

    if sent_count > 0:
        messages.success(request, f'Đã gửi báo cáo thành công cho {sent_count} cố vấn học tập.')
    if errors:
        messages.error(request, f'Lỗi khi gửi báo cáo cho {len(errors)} cố vấn: {", ".join(errors[:3])}')

    return redirect(f"/dashboard/bao-cao/?nam_hoc={nam_hoc_filter}&ky={ky_filter}")
