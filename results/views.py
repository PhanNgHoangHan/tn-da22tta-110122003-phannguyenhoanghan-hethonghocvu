import csv
import io
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
import openpyxl
from .models import KetQuaHocTap
from .forms import KetQuaForm, ImportDiemForm, FilterKetQuaForm
from .utils import tinh_dtbctl, tinh_dtbchk, get_phan_phoi_diem, kiem_tra_canh_bao, dem_canh_bao_lien_tiep, xac_dinh_muc_canh_bao, dong_bo_canh_bao_sinh_vien, la_gdtc
from students.models import SinhVien, MonHoc, HocKy
from academic_warnings.models import CanhBaoHocVu


def _xu_ly_canh_bao(sv, hoc_ky):
    """Kiểm tra và lưu cảnh báo học vụ sau khi nhập/sửa điểm."""
    dong_bo_canh_bao_sinh_vien(sv)


@login_required
def ketqua_list(request):
    from collections import OrderedDict
    from results.utils import la_dat, diem_he4 as _diem_he4

    # --- Lọc cơ bản: chỉ lấy lần học 1 để hiển thị ---
    qs = KetQuaHocTap.objects.select_related(
        'sinh_vien', 'sinh_vien__lop', 'mon_hoc', 'hoc_ky'
    ).filter(lan_hoc=1)

    if request.user.is_sinhvien:
        sv = SinhVien.objects.filter(user=request.user).first()
        qs = qs.filter(sinh_vien=sv) if sv else qs.none()
    elif request.user.is_covan:
        qs = qs.filter(sinh_vien__lop__covan=request.user)

    # Lọc học kỳ (dùng chung cho cả 2 chế độ)
    hoc_ky_filter = request.GET.get('hoc_ky', '')
    if hoc_ky_filter:
        qs = qs.filter(hoc_ky_id=hoc_ky_filter)

    # Lọc điểm hệ 4 (chỉ cho cố vấn/giáo vụ)
    he4_filter = request.GET.get('he4', '')

    # Lọc phân cấp thông minh cho giáo vụ / admin
    khoa = request.GET.get('khoa', '')
    nganh_id = request.GET.get('nganh', '')
    lop_id = request.GET.get('lop', '')
    khoa_hoc = request.GET.get('khoa_hoc', '')

    if not request.user.is_sinhvien:
        # Áp dụng bộ lọc phân cấp cho cả Giáo vụ và Cố vấn
        if khoa_hoc:
            cohort_suffix = khoa_hoc[-2:]
            qs = qs.filter(sinh_vien__lop__ten_lop__istartswith=f'DA{cohort_suffix}')
        if khoa:
            qs = qs.filter(sinh_vien__nganh__khoa=khoa)
        if nganh_id:
            qs = qs.filter(sinh_vien__nganh_id=nganh_id)
        if lop_id:
            qs = qs.filter(sinh_vien__lop_id=lop_id)

    hockys = HocKy.objects.all().order_by('-nam_hoc', '-ky')

    # ================================================================
    # CHẾ ĐỘ SINH VIÊN: hiển thị điểm từng môn theo năm → HK (giảm dần)
    # ================================================================
    if request.user.is_sinhvien:
        # Nhóm theo năm (giảm dần) → HK → [kết quả]
        theo_nam = OrderedDict()
        for kq in qs.order_by('-hoc_ky__nam_hoc', '-hoc_ky__ky', 'mon_hoc__ten_mh'):
            nam = kq.hoc_ky.nam_hoc
            hk_ten = str(kq.hoc_ky)
            if nam not in theo_nam:
                theo_nam[nam] = OrderedDict()
            if hk_ten not in theo_nam[nam]:
                theo_nam[nam][hk_ten] = {'ket_qua': [], 'dtbchk_10': 0.0, 'dtbchk_4': 0.0,
                                          'tc_dang_ky': 0, 'tc_dat': 0}
            theo_nam[nam][hk_ten]['ket_qua'].append(kq)

        # Tính ĐTBCHK từng HK và tổng kết từng năm (loại bỏ thể chất khỏi ĐTB, hiển thị tích lũy cộng dồn)
        for nam, hk_dict in theo_nam.items():
            for hk_ten, data in hk_dict.items():
                kqs = data['ket_qua']
                tc_dang_ky = sum(k.mon_hoc.so_tc for k in kqs)
                
                # Tính ĐTBCHK (loại bỏ Giáo dục thể chất)
                tong_tc_gpa = sum(k.mon_hoc.so_tc for k in kqs if k.diem_tk is not None and not la_gdtc(k.mon_hoc.ma_mh))
                if tong_tc_gpa > 0:
                    d10 = round(sum(k.diem_tk * k.mon_hoc.so_tc for k in kqs if k.diem_tk is not None and not la_gdtc(k.mon_hoc.ma_mh)) / tong_tc_gpa, 2)
                    d4  = round(sum((_diem_he4(k.diem_tk) or 0) * k.mon_hoc.so_tc for k in kqs if k.diem_tk is not None and not la_gdtc(k.mon_hoc.ma_mh)) / tong_tc_gpa, 2)
                else:
                    d10 = d4 = 0.0
                    
                data['dtbchk_10']  = d10
                data['dtbchk_4']   = d4
                data['tc_dang_ky'] = tc_dang_ky
                data['tc_dat']     = sum(k.mon_hoc.so_tc for k in kqs if la_dat(k.diem_tk))
                
                # Tính điểm tích lũy tính đến học kỳ này
                hk_obj = kqs[0].hoc_ky if kqs else None
                if hk_obj and sv:
                    ctl_10, ctl_4, tc_tl_val, _ = tinh_dtbctl(sv, hk_obj)
                else:
                    ctl_10 = ctl_4 = 0.0
                    tc_tl_val = 0
                
                data['dtbctl_10'] = ctl_10
                data['dtbctl_4']  = ctl_4
                data['tc_tl']     = tc_tl_val
            

        # Tổng kết toàn khóa (sử dụng tinh_dtbctl để lấy điểm tích lũy chính xác)
        if sv:
            dtbctl_10, dtbctl_4, tc_tl, _ = tinh_dtbctl(sv)
            toan_khoa = {
                'dtb_10': dtbctl_10,
                'dtb_4':  dtbctl_4,
                'tc_tl':  tc_tl,
            }
        else:
            toan_khoa = {
                'dtb_10': 0.0,
                'dtb_4':  0.0,
                'tc_tl':  0,
            }

        return render(request, 'results/ketqua_list.html', {
            'theo_nam_sv': theo_nam,
            'toan_khoa': toan_khoa,
            'hockys': hockys,
            'hoc_ky_filter': hoc_ky_filter,
            'is_sinhvien': True,
        })

    # ================================================================
    # CHẾ ĐỘ CỐ VẤN / GIÁO VỤ / ADMIN:
    # Hiển thị tổng hợp theo lớp → năm (giảm dần) → HK → sinh viên + điểm HK
    # ================================================================

    # Tính ĐTBCHK cho từng (sinh_vien, hoc_ky)
    # Nhóm: lớp → năm → HK → sv → [kết quả]
    raw = OrderedDict()
    for kq in qs.order_by(
        'sinh_vien__lop__ten_lop', '-hoc_ky__nam_hoc', '-hoc_ky__ky', 'sinh_vien__mssv'
    ):
        ten_lop = kq.sinh_vien.lop.ten_lop if kq.sinh_vien.lop else 'Chưa xếp lớp'
        nam = kq.hoc_ky.nam_hoc
        hk_ten = str(kq.hoc_ky)
        sv_key = kq.sinh_vien.pk

        raw.setdefault(ten_lop, OrderedDict())
        raw[ten_lop].setdefault(nam, OrderedDict())
        raw[ten_lop][nam].setdefault(hk_ten, OrderedDict())
        raw[ten_lop][nam][hk_ten].setdefault(sv_key, {
            'sv': kq.sinh_vien, 'ket_qua': []
        })
        raw[ten_lop][nam][hk_ten][sv_key]['ket_qua'].append(kq)

    # Tính ĐTBCHK hệ 10 và hệ 4 cho từng SV-HK, áp dụng lọc he4
    theo_lop = OrderedDict()
    for ten_lop, theo_nam in raw.items():
        for nam, theo_hk in theo_nam.items():
            for hk_ten, sv_dict in theo_hk.items():
                rows = []
                for sv_key, data in sv_dict.items():
                    kqs = data['ket_qua']
                    tong_tc = sum(k.mon_hoc.so_tc for k in kqs if k.diem_tk is not None)
                    if tong_tc == 0:
                        dtbchk_10 = dtbchk_4 = 0.0
                    else:
                        dtbchk_10 = round(sum(k.diem_tk * k.mon_hoc.so_tc for k in kqs if k.diem_tk is not None) / tong_tc, 2)
                        dtbchk_4  = round(sum((_diem_he4(k.diem_tk) or 0) * k.mon_hoc.so_tc for k in kqs if k.diem_tk is not None) / tong_tc, 2)
                    tc_dat = sum(k.mon_hoc.so_tc for k in kqs if la_dat(k.diem_tk))

                    # Áp dụng lọc điểm hệ 4
                    if he4_filter:
                        try:
                            he4_val = float(he4_filter)
                            if dtbchk_4 < he4_val:
                                continue
                        except ValueError:
                            pass

                    rows.append({
                        'sv': data['sv'],
                        'dtbchk_10': dtbchk_10,
                        'dtbchk_4': dtbchk_4,
                        'tc_dang_ky': tong_tc,
                        'tc_dat': tc_dat,
                        'so_mon': len(kqs),
                    })

                if not rows:
                    continue
                theo_lop.setdefault(ten_lop, OrderedDict())
                theo_lop[ten_lop].setdefault(nam, OrderedDict())
                theo_lop[ten_lop][nam][hk_ten] = rows

    # Dữ liệu cho bộ lọc phân cấp (để hiển thị dropdown)
    from students.models import Nganh, Lop
    import re
    
    nganhs = Nganh.objects.exclude(ma_nganh__iexact='None').exclude(ten_nganh__iexact='None')
    khoas = [k for k in Nganh.objects.exclude(khoa__isnull=True).exclude(khoa='').values_list('khoa', flat=True).distinct() if k.lower() != 'none']
    
    # Lấy danh sách khóa học duy nhất từ các lớp
    unique_years = set()
    for name in Lop.objects.values_list('ten_lop', flat=True):
        if name:
            m = re.search(r'\d{2}', name)
            if m:
                unique_years.add("20" + m.group())
    khoa_hocs = sorted(list(unique_years), reverse=True)

    lops = Lop.objects.exclude(ten_lop__iexact='None')
    if request.user.is_covan:
        lops = lops.filter(covan=request.user)
    if nganh_id:
        lops = lops.filter(nganh_id=nganh_id)
    if khoa_hoc:
        cohort_suffix = khoa_hoc[-2:]
        lops = lops.filter(ten_lop__istartswith=f'DA{cohort_suffix}')
        
    if khoa:
        nganhs = nganhs.filter(khoa=khoa)
        lops = lops.filter(nganh__khoa=khoa)

    return render(request, 'results/ketqua_list.html', {
        'theo_lop': theo_lop,
        'hockys': hockys,
        'hoc_ky_filter': hoc_ky_filter,
        'he4_filter': he4_filter,
        'is_sinhvien': False,
        'nganhs': nganhs,
        'khoas': khoas,
        'lops': lops,
        'khoa_hocs': khoa_hocs,
        'selected_khoa': khoa,
        'nganh_id': nganh_id,
        'selected_lop': lop_id,
        'selected_khoa_hoc': khoa_hoc,
    })


def _nhom_theo_nam_hk(qs):
    """Nhóm queryset theo năm học giảm dần → học kỳ."""
    from collections import OrderedDict
    result = OrderedDict()
    for kq in qs.order_by('-hoc_ky__nam_hoc', '-hoc_ky__ky', 'mon_hoc__ten_mh'):
        nam = kq.hoc_ky.nam_hoc
        hk_ten = str(kq.hoc_ky)
        if nam not in result:
            result[nam] = OrderedDict()
        if hk_ten not in result[nam]:
            result[nam][hk_ten] = []
        result[nam][hk_ten].append(kq)
    return result

@login_required
def ketqua_create(request):
    if request.user.is_sinhvien or request.user.is_covan:
        messages.error(request, 'Bạn không có quyền nhập điểm.')
        return redirect('results:ketqua_list')
    form = KetQuaForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        kq = form.save()
        _xu_ly_canh_bao(kq.sinh_vien, kq.hoc_ky)
        messages.success(request, 'Nhập điểm thành công.')
        return redirect('results:ketqua_list')
    return render(request, 'results/ketqua_form.html', {'form': form, 'title': 'Nhập điểm'})


@login_required
def ketqua_edit(request, pk):
    if request.user.is_sinhvien or request.user.is_covan:
        messages.error(request, 'Bạn không có quyền sửa điểm.')
        return redirect('results:ketqua_list')
    kq = get_object_or_404(KetQuaHocTap, pk=pk)
    form = KetQuaForm(request.POST or None, instance=kq)
    if request.method == 'POST' and form.is_valid():
        kq = form.save()
        _xu_ly_canh_bao(kq.sinh_vien, kq.hoc_ky)
        messages.success(request, 'Cập nhật điểm thành công.')
        return redirect('results:ketqua_list')
    return render(request, 'results/ketqua_form.html', {'form': form, 'title': 'Sửa điểm'})


@login_required
def ketqua_delete(request, pk):
    if request.user.is_sinhvien or request.user.is_covan:
        messages.error(request, 'Bạn không có quyền xóa điểm.')
        return redirect('results:ketqua_list')
    kq = get_object_or_404(KetQuaHocTap, pk=pk)
    if request.method == 'POST':
        sv = kq.sinh_vien
        kq.delete()
        _xu_ly_canh_bao(sv, None)
        messages.success(request, 'Xóa kết quả thành công.')
        return redirect('results:ketqua_list')
    return render(request, 'students/confirm_delete.html', {'obj': kq, 'title': 'Xóa kết quả'})


def parse_hoc_ky(value):
    import re
    if not value:
        return None
    val_str = str(value).strip().lower()
    nam_hoc_match = re.search(r'\d{4}-\d{4}', val_str)
    if not nam_hoc_match:
        return None
    nam_hoc = nam_hoc_match.group()
    
    if 'hè' in val_str or 'he' in val_str or 'hk3' in val_str or 'hk 3' in val_str or 'kỳ 3' in val_str or 'ky 3' in val_str:
        ky = '3'
    elif '2' in val_str:
        ky = '2'
    elif '1' in val_str:
        ky = '1'
    else:
        ky = '1'
    return ky, nam_hoc


@login_required
def import_diem(request):
    if request.user.is_sinhvien or request.user.is_covan:
        messages.error(request, 'Bạn không có quyền import điểm.')
        return redirect('results:ketqua_list')
    form = ImportDiemForm(request.POST or None, request.FILES or None)
    errors = []
    if request.method == 'POST' and form.is_valid():
        f = request.FILES['file']
        count = 0
        imported_hk_set = set()
        try:
            if f.name.lower().endswith('.csv'):
                decoded = f.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded))
                rows = list(reader)
            else:
                wb = openpyxl.load_workbook(f)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                rows = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]

            for row in rows:
                if not row:
                    continue
                mssv_val = row.get('mssv')
                ma_mh_val = row.get('ma_mh')
                if mssv_val is None or ma_mh_val is None:
                    continue
                mssv = str(mssv_val).strip()
                ma_mh = str(ma_mh_val).strip()
                if not mssv or mssv.lower() == 'none' or not ma_mh or ma_mh.lower() == 'none':
                    continue
                try:
                    # Lấy hoặc tự động tạo học kỳ
                    hoc_ky_obj = None
                    ky_val = row.get('ky') or row.get('kỳ') or row.get('ky_hoc')
                    nam_hoc_val = row.get('nam_hoc') or row.get('năm học')
                    
                    if ky_val is not None and nam_hoc_val is not None:
                        ky = str(ky_val).strip()
                        nam_hoc = str(nam_hoc_val).strip()
                        if ky in ('1', '2', '3'):
                            pass
                        elif 'hè' in ky.lower() or 'he' in ky.lower() or '3' in ky:
                            ky = '3'
                        elif '2' in ky:
                            ky = '2'
                        else:
                            ky = '1'
                        
                        if nam_hoc:
                            hoc_ky_obj, _ = HocKy.objects.get_or_create(ky=ky, nam_hoc=nam_hoc)
                            
                    if hoc_ky_obj is None:
                        hoc_ky_str = row.get('hoc_ky') or row.get('học kỳ') or row.get('học kì') or row.get('hoc_ki')
                        if hoc_ky_str:
                            res = parse_hoc_ky(hoc_ky_str)
                            if res:
                                ky, nam_hoc = res
                                hoc_ky_obj, _ = HocKy.objects.get_or_create(ky=ky, nam_hoc=nam_hoc)
                                
                    if hoc_ky_obj is None:
                        raise Exception("Không xác định được học kỳ (cột 'hoc_ky' hoặc cột 'ky'/'nam_hoc')")

                    sv = SinhVien.objects.get(mssv=mssv)
                    mh = MonHoc.objects.get(ma_mh=ma_mh)
                    diem_qt  = float(row['diem_qt'])  if row.get('diem_qt')  not in (None, '', 'None') else None
                    diem_thi = float(row['diem_thi']) if row.get('diem_thi') not in (None, '', 'None') else None
                    diem_tk  = float(row['diem_tk'])  if row.get('diem_tk')  not in (None, '', 'None') else None
                    
                    KetQuaHocTap.objects.update_or_create(
                        sinh_vien=sv, mon_hoc=mh, hoc_ky=hoc_ky_obj,
                        lan_hoc=int(row.get('lan_hoc', 1) or 1),
                        defaults={'diem_qt': diem_qt, 'diem_thi': diem_thi, 'diem_tk': diem_tk}
                    )
                    imported_hk_set.add(hoc_ky_obj)
                    count += 1
                except Exception as e:
                    errors.append(f"MSSV {mssv}, Mã MH {ma_mh}: {e}")

            # Kiểm tra cảnh báo cho tất cả SV trong những học kỳ đã import
            for hk in imported_hk_set:
                for sv in SinhVien.objects.filter(ket_qua__hoc_ky=hk).distinct():
                    _xu_ly_canh_bao(sv, hk)

            messages.success(request, f'Import thành công {count} kết quả.')
            if errors:
                messages.warning(request, f'{len(errors)} lỗi: ' + '; '.join(errors[:5]))
            return redirect('results:ketqua_list')
        except Exception as e:
            messages.error(request, f'Lỗi đọc file: {e}')
    return render(request, 'results/import_diem.html', {'form': form, 'errors': errors})


@login_required
def export_diem(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Kết quả học tập'
    ws.append(['MSSV', 'Họ tên', 'Môn học', 'Số TC', 'Học kỳ',
               'Điểm QT', 'Điểm thi', 'Điểm TK', 'Điểm chữ', 'Hệ 4', 'Lần học'])
    qs = KetQuaHocTap.objects.select_related('sinh_vien', 'mon_hoc', 'hoc_ky').all()
    if request.user.is_sinhvien:
        qs = qs.filter(sinh_vien__user=request.user)
    elif request.user.is_covan:
        qs = qs.filter(sinh_vien__lop__covan=request.user)
    for kq in qs:
        ws.append([kq.sinh_vien.mssv, kq.sinh_vien.ho_ten, kq.mon_hoc.ten_mh,
                   kq.mon_hoc.so_tc, str(kq.hoc_ky),
                   kq.diem_qt, kq.diem_thi, kq.diem_tk, kq.diem_chu, kq.diem_he4, kq.lan_hoc])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ket_qua_hoc_tap.xlsx"'
    wb.save(response)
    return response


@login_required
def api_gpa_chart(request, sv_id):
    sv = get_object_or_404(SinhVien, pk=sv_id)
    hockys = HocKy.objects.filter(ket_qua__sinh_vien=sv).distinct().order_by('nam_hoc', 'ky')
    labels, gpas = [], []
    for hk in hockys:
        dtbchk, tc = tinh_dtbchk(sv, hk)
        if tc > 0:
            labels.append(str(hk))
            gpas.append(dtbchk)
    return JsonResponse({'labels': labels, 'gpas': gpas})
